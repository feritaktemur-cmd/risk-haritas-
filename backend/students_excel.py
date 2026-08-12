"""Student Excel PREVIEW analysis — READ ONLY.

NEVER writes to the database. Given the uploaded .xlsx bytes plus the
logged-in school's context (preschool flag, defined classes, existing
student numbers), it validates each row and returns a preview payload.
No INSERT/UPDATE/DELETE. Pure function (context injected) so it needs no
DB access itself.

Fixed, standard columns (the user prepares the file to this format):
    Öğrenci No | Ad | Soyad | Sınıf | Şube
No E-Okul auto-mapping, no "Ad Soyad" single-column support.
"""
import io
import re

from openpyxl import load_workbook

from excel_preview import norm  # reuse Turkish-aware normalization

ST_READY = "Hazır"
ST_ERROR = "Hatalı"


def _find_header(rows):
    """Locate header row + column indexes for the five fixed columns."""
    for r_idx, row in enumerate(rows[:15]):
        cols = {"number": None, "first_name": None, "last_name": None, "level": None, "branch": None}
        for c_idx, cell in enumerate(row):
            h = norm(cell)
            if not h:
                continue
            if cols["last_name"] is None and "soyad" in h:
                cols["last_name"] = c_idx
            elif cols["first_name"] is None and h in ("ad", "adı", "adi"):
                cols["first_name"] = c_idx
            if cols["number"] is None and ("no" in h or "numara" in h):
                cols["number"] = c_idx
            if cols["level"] is None and ("sınıf" in h or "sinif" in h):
                cols["level"] = c_idx
            if cols["branch"] is None and ("şube" in h or "sube" in h):
                cols["branch"] = c_idx
        if all(v is not None for v in cols.values()):
            return r_idx, cols
    return None, None


def _cell_text(value) -> str:
    """Stringify a cell preserving leading zeros for text-entered values.

    Text cells (e.g. '0012') come back as str and are preserved as-is.
    Numeric cells come back as int/float; integer-valued numbers are shown
    without a trailing '.0'.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    return str(value).strip()


def analyze_student_rows(file_bytes: bytes, is_preschool: bool, defined_classes: set, existing_numbers: set) -> dict:
    """Pure preview analysis. No DB access.

    defined_classes: set of (level:int, branch:str) already defined by the school.
    existing_numbers: set of student_number strings already in the school.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    header_idx, cols = _find_header(all_rows)
    if cols is None:
        return {
            "error": (
                "Excel başlıkları bulunamadı. Sırasıyla şu sütunlar gereklidir: "
                "'Öğrenci No', 'Ad', 'Soyad', 'Sınıf', 'Şube'."
            ),
        }

    allowed_levels = {4, 5} if is_preschool else set(range(1, 13))

    data_rows = all_rows[header_idx + 1:]
    result_rows = []
    summary = {"total": 0, "valid": 0, "invalid": 0}
    seen_numbers = {}  # normalized number -> first row index (for in-file dup)

    for row in data_rows:
        def cell(k):
            i = cols[k]
            return row[i] if (i is not None and i < len(row)) else None

        raw_number = cell("number")
        raw_first = cell("first_name")
        raw_last = cell("last_name")
        raw_level = cell("level")
        raw_branch = cell("branch")

        # Skip fully empty rows.
        if not any(norm(x) for x in (raw_number, raw_first, raw_last, raw_level, raw_branch)):
            continue

        summary["total"] += 1

        number_s = _cell_text(raw_number).strip()
        first_s = _cell_text(raw_first).strip()
        last_s = _cell_text(raw_last).strip()

        # Branch: single Latin letter, normalized to uppercase.
        branch_raw = _cell_text(raw_branch).strip()
        branch_s = branch_raw.upper()

        # Level display + parse.
        level_display = _cell_text(raw_level).strip()
        level_int = None
        if level_display:
            try:
                level_int = int(float(level_display))
            except (TypeError, ValueError):
                level_int = None

        errors = []

        if not number_s:
            errors.append("Öğrenci no boş olamaz.")
        if not first_s:
            errors.append("Ad boş olamaz.")
        if not last_s:
            errors.append("Soyad boş olamaz.")

        if level_int is None or level_int not in allowed_levels:
            if is_preschool:
                errors.append("Anaokulunda yaş grubu yalnızca 4 veya 5 olabilir.")
            else:
                errors.append("Sınıf 1 ile 12 arasında bir sayı olmalıdır.")

        if not re.fullmatch(r"[A-Z]", branch_s):
            errors.append("Şube yalnızca tek bir büyük harf (A-Z) olabilir.")

        # Class must be defined in the school (only when level+branch parse OK).
        if level_int is not None and re.fullmatch(r"[A-Z]", branch_s):
            if (level_int, branch_s) not in defined_classes:
                errors.append("Bu sınıf okulda tanımlı değil.")

        # Already in DB for this school.
        if number_s and number_s in existing_numbers:
            errors.append("Bu öğrenci numarası zaten kayıtlı.")

        # In-file duplicate.
        if number_s:
            nkey = number_s
            if nkey in seen_numbers:
                errors.append("Bu öğrenci numarası dosyada birden fazla kez var.")
            else:
                seen_numbers[nkey] = summary["total"]

        status = ST_READY if not errors else ST_ERROR
        if errors:
            summary["invalid"] += 1
        else:
            summary["valid"] += 1

        result_rows.append({
            "student_number": number_s,
            "first_name": first_s,
            "last_name": last_s,
            "level": level_display,
            "branch": branch_raw,
            "status": status,
            "error": "; ".join(errors) if errors else "",
        })

    return {"summary": summary, "rows": result_rows}
