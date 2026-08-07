"""Excel PREVIEW analysis — READ ONLY.

This module NEVER writes to the database. It only reads reference data
(districts, school_types) and analyzes an uploaded Excel file to produce a
preview. No INSERT/UPDATE/DELETE. No school records are created.

`analyze_rows` is pure (reference data injected) so it can be unit-tested
without any database access.
"""
import io
import re

from openpyxl import load_workbook

from supabase_client import get_service_client

# --- Turkish-aware normalization ------------------------------------------
_TR_LOWER = str.maketrans({
    "I": "ı", "İ": "i", "Ş": "ş", "Ğ": "ğ", "Ü": "ü", "Ö": "ö", "Ç": "ç",
})


def norm(value) -> str:
    """Trim, collapse whitespace, Turkish-aware lowercase.

    Makes spacing and minor casing differences irrelevant when matching.
    """
    if value is None:
        return ""
    s = re.sub(r"\s+", " ", str(value).strip())
    return s.translate(_TR_LOWER).lower()


# --- Business rules --------------------------------------------------------
# Institutions that are automatically OUT OF SCOPE (matched by substring).
OUT_OF_SCOPE = [
    "Rehberlik ve Araştırma Merkezi",   # RAM
    "Bilim ve Sanat Merkezi",           # BİLSEM
    "İlçe Milli Eğitim Müdürlüğü",
    "Halk Eğitimi Merkezi",
]
_OUT_OF_SCOPE_N = [norm(x) for x in OUT_OF_SCOPE]

# MEB name -> system school_types name aliases.
MEB_ALIASES = {
    norm("Anadolu Meslek Programı"): "Mesleki ve Teknik Anadolu Lisesi",
    norm("Özel Eğitim Meslek Okulu (Zihinsel Engelliler)"): "Özel Eğitim Meslek Okulu",
}

# Status labels
ST_LOADABLE = "YÜKLENEBİLİR"
ST_OUT = "KAPSAM DIŞI"
ST_BAD_DISTRICT = "HATALI İLÇE"
ST_BAD_TYPE = "HATALI OKUL TÜRÜ"

VALID_MANAGEMENT_TYPES = {"Resmî", "Özel"}


def _find_header(rows):
    """Locate header row + column indexes for name / district / MEB type."""
    for r_idx, row in enumerate(rows[:15]):
        cols = {"name": None, "district": None, "type": None}
        for c_idx, cell in enumerate(row):
            h = norm(cell)
            if not h:
                continue
            if cols["name"] is None and ("kurum ad" in h or "okul ad" in h or h in ("ad", "adı")):
                cols["name"] = c_idx
            if cols["district"] is None and ("ilçe" in h or "ilce" in h):
                cols["district"] = c_idx
            if cols["type"] is None and ("tür" in h or "tur" in h):
                cols["type"] = c_idx
        if all(v is not None for v in cols.values()):
            return r_idx, cols
    return None, None


def _map_school_type(meb_n, st_index):
    """Map a normalized MEB type name to a system school_type dict or None."""
    if meb_n in MEB_ALIASES:
        target = norm(MEB_ALIASES[meb_n])
        if target in st_index:
            return st_index[target]
    if meb_n in st_index:
        return st_index[meb_n]
    return None


def analyze_rows(file_bytes: bytes, management_type: str, districts: list, school_types: list) -> dict:
    """Pure analysis. No DB access. Returns the preview payload."""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    header_idx, cols = _find_header(all_rows)
    if cols is None:
        return {
            "error": "Excel başlıkları bulunamadı. 'Kurum Adı', 'İlçe' ve 'Kurum Türü' sütunları gereklidir.",
        }

    # Active reference indexes (normalized)
    district_index = {norm(d["name"]): d for d in districts if d.get("is_active", True)}
    st_index = {norm(s["name"]): s for s in school_types if s.get("is_active", True)}

    data_rows = all_rows[header_idx + 1:]
    result_rows = []
    summary = {"total": 0, "loadable": 0, "out_of_scope": 0, "invalid_district": 0, "invalid_school_type": 0}

    for row in data_rows:
        name = row[cols["name"]] if cols["name"] < len(row) else None
        district = row[cols["district"]] if cols["district"] < len(row) else None
        meb = row[cols["type"]] if cols["type"] < len(row) else None

        # Skip fully empty rows
        if not any(norm(x) for x in (name, district, meb)):
            continue

        summary["total"] += 1
        name_s = str(name).strip() if name is not None else ""
        district_s = str(district).strip() if district is not None else ""
        meb_s = str(meb).strip() if meb is not None else ""
        meb_n = norm(meb)
        district_n = norm(district)

        system_type = "-"
        if any(k in meb_n for k in _OUT_OF_SCOPE_N):
            status = ST_OUT
            summary["out_of_scope"] += 1
        else:
            district_ok = district_n in district_index
            mapped = _map_school_type(meb_n, st_index)
            if not district_ok:
                status = ST_BAD_DISTRICT
                summary["invalid_district"] += 1
            elif mapped is None:
                status = ST_BAD_TYPE
                summary["invalid_school_type"] += 1
            else:
                status = ST_LOADABLE
                system_type = mapped["name"]
                summary["loadable"] += 1

        result_rows.append({
            "institution_name": name_s,
            "district": district_s,
            "meb_type": meb_s,
            "system_school_type": system_type,
            "status": status,
        })

    return {
        "management_type": management_type,
        "summary": summary,
        "rows": result_rows,
    }


def load_reference():
    """READ-ONLY fetch of districts & school_types from Supabase."""
    client = get_service_client()
    districts = client.table("districts").select("id,name,is_active").execute().data
    school_types = client.table("school_types").select("id,name,is_active").execute().data
    return districts, school_types
